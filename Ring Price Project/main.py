import sys
import os
import time
import requests
import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from selenium import webdriver
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from webdriver_manager.firefox import GeckoDriverManager
from modules.database import init_db, save_to_sql

# Import your custom modules
sys.path.append(os.path.abspath(os.path.join('..')))
import modules.scraper as scraper
import modules.extractor as extractor

try:
    from config.config import LOGIN_EMAIL, LOGIN_PASSWORD, main_url
except ImportError:
    LOGIN_EMAIL = os.getenv('LOGIN_EMAIL')
    LOGIN_PASSWORD = os.getenv('LOGIN_PASSWORD')
    main_url = os.getenv('MAIN_URL', 'https://example.com/login')

try:
    from config.model_urls import model_urls
    print("Successfully loaded local URL configuration.")
except ImportError:
    model_urls = []
    print("-" * 50)
    print("NOTE: For data privacy and to respect source integrity, the URL list has been hidden.")
    print("To test this functionality, please provide your URLs in 'config/model_urls.py' or contact me for my version.")


# --- 1. SETUP FIREFOX BROWSER ---
def init_driver():
    """Initializes the Firefox WebDriver with anti-detection settings."""
    options = FirefoxOptions()
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.set_preference("general.useragent.override", "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0")
    options.set_preference("dom.webdriver.enabled", False)
    
    service = FirefoxService(GeckoDriverManager().install())
    driver = webdriver.Firefox(service=service, options=options)
    driver.set_window_size(1920, 1080)
    return driver

# --- 2. EXCEL SAVING LOGIC (Maintains Image support for the exam) ---
def save_results_to_xlsx(all_details):
    """
    Saves extracted data to Excel and embeds product images in the first column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Data"

    # Define headers
    headers = ['Image', 'Metal', 'Refcode', 'Description', 'Price']
    ws.append(headers)

    for row_idx, item in enumerate(all_details, start=2):
        # 1. Fill Text Data
        ws.cell(row=row_idx, column=2, value=item.get('Metal'))
        ws.cell(row=row_idx, column=3, value=item.get('Refcode'))
        ws.cell(row=row_idx, column=4, value=item.get('Description'))
        ws.cell(row=row_idx, column=5, value=item.get('Price'))

        # 2. Handle Image Download and Insertion
        img_url = item.get('Image URL')
        if img_url:
            try:
                response = requests.get(img_url, timeout=10)
                if response.status_code == 200:
                    img_data = BytesIO(response.content)
                    img = OpenpyxlImage(img_data)
                    
                    # Resize image to fit cell (approx 80x80 pixels)
                    img.width = 80
                    img.height = 80
                    
                    # Anchoring image to Column A
                    ws.add_image(img, f'A{row_idx}')
                    # Set row height to accommodate image
                    ws.row_dimensions[row_idx].height = 65 
            except Exception as e:
                print(f"Could not download image for {item.get('Refcode')}: {e}")

    # Column Formatting
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['D'].width = 40 # Description usually long
    
    filename = os.path.join('data', 'raw', 'ring_details.xlsx')
    wb.save(filename)
    print(f"Excel file '{filename}' generated successfully!")

# --- 3. MAIN EXECUTION FLOW ---
def main():
    if not LOGIN_EMAIL or not LOGIN_PASSWORD:
        print("Error: Missing credentials!")
        print("Please set config/config.py or environment variables.")
        sys.exit(1)
        
    if "example.com" in main_url:
        print("Warning: Using placeholder URL. The scraper might not work as expected.")

    os.makedirs(os.path.join('data', 'raw'), exist_ok=True)
    os.makedirs(os.path.join('data', 'processed'), exist_ok=True)
    
    driver = init_driver()
    # Load URLs (Replace with your testing_urls or list)

    driver = init_driver()
    all_extracted_data = []
    init_db()

    try:
        # Step A: Perform Login
        print("Starting login process...")
        if scraper.perform_login(driver, LOGIN_EMAIL, LOGIN_PASSWORD, main_url):
            
            # Step B: Loop through URLs
            for i, url in enumerate(model_urls):
                print(f"Processing ({i+1}/{len(model_urls)}): {url}")
                
                # Use Scraper Module
                html = scraper.fetch_page_source(driver, url)
                
                if html:
                    # Use Extractor Module
                    data = extractor.extract_ring_details(html)
                    if data:
                        all_extracted_data.extend(data)
                        print(f"Successfully extracted {len(data)} variations.")
                
                time.sleep(1) # Small delay between pages

        # Step C: Save all data to Excel
        if all_extracted_data:
            save_results_to_xlsx(all_extracted_data)
            save_to_sql(all_extracted_data)
        else:
            print("No data was extracted.")

    finally:
        driver.quit()
        print("Browser closed. Task finished.")

# Run the script
if __name__ == "__main__":
    main()